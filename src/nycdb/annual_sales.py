import re
import openpyxl # for xlsx files
import xlrd     # for xls files
from datetime import datetime
from os.path import basename, splitext

headers = [
    'borough',
    'neighborhood',
    'building_class_category',
    'tax_class_at_present',
    'block',
    'lot',
    'easement',
    'building_class_at_present',
    'address',
    'apartment_number',
    'zip_code',
    'residential_units',
    'commercial_units',
    'total_units',
    'land_square_feet',
    'gross_square_feet',
    'year_built',
    'tax_class_at_time_of_sale',
    'building_class_at_time_of_sale',
    'sale_price',
    'sale_date',
    'year'
]

class AnnualSales:
    def __init__(self, filename):
        self.filename = filename
        self.year = re.search(r"\d{4}", basename(self.filename)).group()
        self.ext = splitext(self.filename)[1]

    def __iter__(self):
        for row in self.iter_rows():
            if row[0] in ['1', '2', '3', '4', '5']:
                yield dict(zip(headers, row + [self.year]))

    def iter_rows(self):
        if self.ext == '.xlsx':
            workbook = openpyxl.load_workbook(self.filename)
            worksheet = workbook[workbook.sheetnames[0]]
            for row in worksheet.iter_rows(values_only=True):
                yield list(row)
        else:
            workbook = xlrd.open_workbook(self.filename)
            worksheet = workbook.sheet_by_index(0)

            for row in worksheet.get_rows():
                values = [ x.value for x in row ]

                try:
                    values[0] = str(int(values[0]))
                except ValueError:
                    continue

                if values[0] not in ['1', '2', '3', '4', '5']:
                    continue

                values[2] = values[2].strip()

                if isinstance(values[3], float):
                    values[3] = str(int(values[3])) # tax class
                values[4] = str(int(values[4])) # block
                values[5] = str(int(values[5])) # lot
                values[6] = values[6].strip()

                if isinstance(values[9], float):
                    values[9] = str(int(values[9]))

                values[10] = str(int(values[10])).zfill(5)

                # res units, com units, land sqft, gross sqft, year build
                for i in range(11,17):
                    if values[i] != None:
                        values[i] = int(values[i])

                values[17] = str(int(values[17]))
                values[19] = str(int(values[19])) # sale price
                values[20] = datetime(*xlrd.xldate_as_tuple(values[20], workbook.datemode))

                yield values
