import re
import os
import openpyxl
from .utility import consume

HEADERS = ['borough', 'neighborhood', 'building_class_category', 'tax_class_at_present', 'block', 'lot', 'building_class_at_present', 'address', 'zip_code', 'residential_units', 'commercial_units', 'total_units', 'land_square_feet', 'gross_square_feet', 'year_built', 'fiscalyear']

def validate_header_row(row):
    for i, s in enumerate(map(lambda s: s.lower().replace('\n', '_').replace(' ', '_'), row)):
        if s == HEADERS[i] or s == HEADERS[i].replace('_at_present', ''):
            continue
        else:
            raise Exception(f"Invalid 421a header formatting for {HEADERS[i]}")


def iter_421a(filename):
    fiscalyear = re.search(r"_(\d{4})_", os.path.basename(filename)).group(1)
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    consume(rows, 3)

    header_row = list(next(rows))

    if header_row[0] is None:
        header_row = list(next(rows))

    validate_header_row(header_row)

    for i, row in enumerate(rows):
        if i == 0 and row[0] is None:
            continue

        row = list(row)
        row.append(fiscalyear)

        if isinstance(row[0], str):
            row[0] = int(row[0])

        if not row[0] in [1, 2, 3, 4, 5]:
            raise Exception(f"Detected bad formatting: {filename}")

        yield dict(zip(HEADERS, row))
